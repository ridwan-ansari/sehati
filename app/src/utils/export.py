import pandas as pd
from io import BytesIO
from sqlalchemy import select, and_
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import datetime, timezone, timedelta
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.src.models.user import User
from app.src.models.user_nutrition import UserNutrition
from app.src.models.exercise_habit import ExerciseHabitAnswer, ExerciseHabitQuestion
from app.src.models.sleep import Sleep
from app.src.models.food import FoodHabitAnswer, FoodHabitQuestion, FoodDiaryAnalysis, FoodDiaryItem, Food


class HealthDataExcelExporter:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.wib = timezone(timedelta(hours=7))

    def _localize(self, dt):
        """Konversi UTC ke WIB"""
        if not dt: 
            return None
        # Jika sudah date object, return as is
        if isinstance(dt, datetime.date) and not isinstance(dt, datetime):
            return dt
        # Konversi datetime ke WIB
        dt_utc = dt.replace(tzinfo=timezone.utc) if dt.tzinfo is None else dt
        return dt_utc.astimezone(self.wib).replace(tzinfo=None)

    def _calculate_age(self, dob):
        """Hitung umur dari tanggal lahir"""
        if not dob: 
            return None
        today = datetime.now().date()
        return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))

    async def get_users_data(self):
        """Ambil data demografi user (exclude admin)"""
        stmt = select(User).where(
            and_(
                User.role != 'admin',
                User.deleted_at.is_(None)
            )
        ).order_by(User.created_at.desc())
        
        result = await self.db.execute(stmt)
        users = result.scalars().all()
        
        return pd.DataFrame([{
            'Fullname': u.fullname,
            'Nickname': u.nickname,
            'Email': u.email,
            'Phone_Number': u.phone_number,
            'Gender': u.gender.value if hasattr(u.gender, 'value') else u.gender,
            'Date_of_Birth': u.date_of_birth,
            'Age': self._calculate_age(u.date_of_birth),
            'Active': 'Yes' if u.active else 'No',
            'Verified': 'Yes' if u.verified else 'No',
            'Created_At': self._localize(u.created_at)
        } for u in users])

    async def get_nutrition_data(self):
        """Ambil data nutrisi dan BMI user"""
        stmt = select(UserNutrition).options(
            selectinload(UserNutrition.user)
        ).where(
            UserNutrition.deleted_at.is_(None)
        ).order_by(UserNutrition.created_at.desc())
        
        result = await self.db.execute(stmt)
        nutritions = result.scalars().all()
        
        data = []
        for n in nutritions:
            if not n.user or n.user.role == 'admin':
                continue
                
            data.append({
                'Nickname': n.user.nickname,
                'Height_cm': n.height_cm,
                'Weight_kg': n.weight_kg,
                'BMI': n.bmi,
                'Ideal_Weight_kg': n.ideal_weight_kg,
                'Status': n.status,
                'Created_At': self._localize(n.created_at),
                'Updated_At': self._localize(n.updated_at) if n.updated_at else None
            })
        
        df = pd.DataFrame(data)
        if not df.empty:
            df = df.sort_values(['Nickname', 'Created_At'], ascending=[True, False])
        return df

    async def get_sleep_records(self):
        """Ambil data pola tidur user"""
        stmt = select(Sleep).options(
            selectinload(Sleep.user)
        ).where(
            Sleep.deleted_at.is_(None)
        ).order_by(Sleep.created_at.desc())
        
        result = await self.db.execute(stmt)
        sleeps = result.scalars().all()
        
        data = []
        for s in sleeps:
            if not s.user or s.user.role == 'admin':
                continue
            
            sleep_time = self._localize(s.sleep_time)
            wake_time = self._localize(s.wake_up_time)
            
            data.append({
                'Nickname': s.user.nickname,
                'Sleep_Time': sleep_time,
                'Wake_Up_Time': wake_time,
                'Sleep_Duration_Minutes': s.sleep_duration_minutes,
                'Target_Sleep_Hours': s.target_sleep_hours,
                'Created_At': self._localize(s.created_at)
            })
        
        df = pd.DataFrame(data)
        if not df.empty:
            df = df.sort_values(['Nickname', 'Sleep_Time'], ascending=[True, False])
        return df

    async def get_food_habit_pivot(self):
        """Ambil data kebiasaan makan dalam format pivot"""
        stmt = select(FoodHabitAnswer).options(
            selectinload(FoodHabitAnswer.user),
            selectinload(FoodHabitAnswer.question)
        ).where(
            FoodHabitAnswer.deleted_at.is_(None)
        ).order_by(FoodHabitAnswer.created_at.desc())
        
        result = await self.db.execute(stmt)
        answers = result.scalars().all()
        
        raw = []
        for a in answers:
            if not (a.user and a.question) or a.user.role == 'admin':
                continue
            
            answer_value = 'Yes' if a.answer else 'No'
            if a.frequency:
                answer_value = f"{answer_value} ({a.frequency})"
            
            raw.append({
                'Nickname': a.user.nickname,
                'Date': self._localize(a.created_at).date(),
                'Question': a.question.question,
                'Answer': answer_value
            })
        
        df = pd.DataFrame(raw)
        if df.empty:
            return df
        
        # Pivot: per user per date, questions jadi kolom
        pivot = df.pivot_table(
            index=['Nickname', 'Date'], 
            columns='Question', 
            values='Answer', 
            aggfunc='first'
        ).reset_index()
        
        return pivot.sort_values(['Nickname', 'Date'], ascending=[True, False])

    async def get_exercise_habit_pivot(self):
        """Ambil data kebiasaan olahraga dalam format pivot"""
        stmt = select(ExerciseHabitAnswer).options(
            selectinload(ExerciseHabitAnswer.user),
            selectinload(ExerciseHabitAnswer.question)
        ).where(
            ExerciseHabitAnswer.deleted_at.is_(None)
        ).order_by(ExerciseHabitAnswer.created_at.desc())
        
        result = await self.db.execute(stmt)
        answers = result.scalars().all()
        
        raw = []
        for a in answers:
            if not (a.user and a.question) or a.user.role == 'admin':
                continue
            
            # Ambil jawaban sesuai tipe
            answer_value = None
            if a.selected_option:
                answer_value = a.selected_option
            elif a.answer_text:
                answer_value = a.answer_text
            else:
                answer_value = '-'
            
            recorded_date = self._localize(a.recorded_at).date() if a.recorded_at else self._localize(a.created_at).date()
            
            raw.append({
                'Nickname': a.user.nickname,
                'Date': recorded_date,
                'Question': a.question.question,
                'Answer': answer_value
            })
        
        df = pd.DataFrame(raw)
        if df.empty:
            return df
        
        # Pivot: per user per date, questions jadi kolom
        pivot = df.pivot_table(
            index=['Nickname', 'Date'], 
            columns='Question', 
            values='Answer', 
            aggfunc='first'
        ).reset_index()
        
        return pivot.sort_values(['Nickname', 'Date'], ascending=[True, False])

    async def get_food_diary_data(self):
        """Ambil data food diary: summary harian dan detail per makanan"""
        stmt = select(FoodDiaryAnalysis).options(
            selectinload(FoodDiaryAnalysis.user),
            selectinload(FoodDiaryAnalysis.items).selectinload(FoodDiaryItem.food)
        ).where(
            FoodDiaryAnalysis.deleted_at.is_(None)
        ).order_by(FoodDiaryAnalysis.created_at.desc())
        
        result = await self.db.execute(stmt)
        analyses = result.unique().scalars().all()
        
        daily_summary = []
        detail_logs = []
        
        for a in analyses:
            if not a.user or a.user.role == 'admin':
                continue
            
            u_name = a.user.nickname
            diary_date = self._localize(a.created_at).date() if a.created_at else None
            
            # Summary harian
            daily_summary.append({
                'Nickname': u_name,
                'Date': diary_date,
                'Activity': a.activity,
                'Energy_Requirement': a.energy_requirement,
                'Desired_Energy_Requirement': a.desired_energy_requirement,
                'Total_Calories': a.total_calories,
                'Reward_Points': a.reward_points,
                'Created_At': self._localize(a.created_at)
            })
            
            # Detail per item makanan
            if a.items:
                for item in a.items:
                    food_name = item.food.name if item.food else 'Unknown'
                    calories = item.food.calories if item.food else 0
                    
                    detail_logs.append({
                        'Nickname': u_name,
                        'Date': diary_date,
                        'Meal_Type': item.meal_type.value if hasattr(item.meal_type, 'value') else item.meal_type,
                        'Food_Name': food_name,
                        'Quantity': item.quantity,
                        'Weight_Grams': item.weight_grams,
                        'Calories_per_100g': calories,
                        'Created_At': self._localize(item.created_at)
                    })
        
        df_daily = pd.DataFrame(daily_summary)
        df_detail = pd.DataFrame(detail_logs)
        
        if not df_daily.empty:
            df_daily = df_daily.sort_values(['Nickname', 'Date'], ascending=[True, False])
        if not df_detail.empty:
            df_detail = df_detail.sort_values(['Nickname', 'Date', 'Meal_Type'], ascending=[True, False, True])
        
        return df_daily, df_detail

    def _apply_styling(self, writer, sheet_name):
        """Apply styling ke Excel sheet"""
        ws = writer.sheets[sheet_name]
        
        # Header styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header styling
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 3, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze first row
        ws.freeze_panes = 'A2'

    async def generate_excel(self) -> BytesIO:
        """Generate Excel file dengan semua data kesehatan"""
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheets_created = []
            
            # Sheet 1: Demographics
            df_users = await self.get_users_data()
            if not df_users.empty:
                df_users.to_excel(writer, sheet_name='1_Demographics', index=False)
                sheets_created.append('1_Demographics')
            
            # Sheet 2: Body Composition & Nutrition
            df_nutrition = await self.get_nutrition_data()
            if not df_nutrition.empty:
                df_nutrition.to_excel(writer, sheet_name='2_Body_Composition', index=False)
                sheets_created.append('2_Body_Composition')
            
            # Sheet 3: Sleep Records
            df_sleep = await self.get_sleep_records()
            if not df_sleep.empty:
                df_sleep.to_excel(writer, sheet_name='3_Sleep_Records', index=False)
                sheets_created.append('3_Sleep_Records')
            
            # Sheet 4: Food Habits
            df_food_habits = await self.get_food_habit_pivot()
            if not df_food_habits.empty:
                df_food_habits.to_excel(writer, sheet_name='4_Food_Habits', index=False)
                sheets_created.append('4_Food_Habits')
            
            # Sheet 5: Exercise Habits
            df_exercise_habits = await self.get_exercise_habit_pivot()
            if not df_exercise_habits.empty:
                df_exercise_habits.to_excel(writer, sheet_name='5_Exercise_Habits', index=False)
                sheets_created.append('5_Exercise_Habits')
            
            # Sheet 6 & 7: Food Diary
            df_diary_daily, df_diary_detail = await self.get_food_diary_data()
            if not df_diary_daily.empty:
                df_diary_daily.to_excel(writer, sheet_name='6_Food_Diary_Summary', index=False)
                sheets_created.append('6_Food_Diary_Summary')
            if not df_diary_detail.empty:
                df_diary_detail.to_excel(writer, sheet_name='7_Food_Diary_Detail', index=False)
                sheets_created.append('7_Food_Diary_Detail')
            
            # Jika tidak ada sheet yang dibuat, buat sheet placeholder
            if not sheets_created:
                df_empty = pd.DataFrame({'Message': ['No data available']})
                df_empty.to_excel(writer, sheet_name='No_Data', index=False)
                sheets_created.append('No_Data')
            
            # Apply styling to all sheets
            for sheet_name in sheets_created:
                self._apply_styling(writer, sheet_name)
        
        output.seek(0)
        return output