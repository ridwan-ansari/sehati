from __future__ import annotations
import re
import hmac
import hashlib
import asyncio
from io import BytesIO
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from loguru import logger
from zoneinfo import ZoneInfo
from datetime import datetime
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, Request, Cookie, UploadFile
from typing import List as FormList

from app.src.core.config import settings
from app.src.router.user.crud import CRUDUser
from app.src.router.food.crud import crud_food
from app.src.router.games.crud import crud_games
from app.src.core.templates import get_templates
from app.src.router.recipe.crud import CRUDRecipe
from app.src.core.session import get_async_session
from app.src.utils.email_client import EmailClient
from app.src.router.merchandise.crud import crud_merch
from app.src.models.video import Video
from app.src.router.video.crud import crud_video
from app.src.models.user_nutrition import UserNutrition
from app.src.utils.export import HealthDataExcelExporter
from app.src.router.appointment.crud import crud_appointment, crud_professional
from app.src.router.merchandise.crud import crud_merch_claim
from app.src.utils.file_service import save_upload_with_uuid
from app.src.router.user_nutrition.crud import CRUDUserNutrition
from app.src.utils.point_service import redeem_merchandise_points
from app.src.router.point.crud import crud_wallet, crud_transaction
from app.src.core.security import Hasher, TokenService, AuthService
from app.src.router.dashboard.blast_crud import crud_blast_log
from app.src.models.blast_log import BlastLog

router = APIRouter()

# In-process real-time progress tracker keyed by blast_id
_blast_progress: dict[str, dict] = {}

crud_user = CRUDUser()
crud_recipe = CRUDRecipe()
templates = get_templates()
email_client = EmailClient()
auth_service = AuthService()
token_service = TokenService()


def _make_csrf_token(admin_token: str) -> str:
    return hmac.new(
        settings.SECRET_KEY.encode("utf-8"),
        admin_token.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()


def _verify_csrf_token(admin_token: str, csrf_token: str) -> bool:
    expected = _make_csrf_token(admin_token)
    return hmac.compare_digest(expected, csrf_token)


def render_page(template, request, **context):
    context.setdefault("year", datetime.now().year)
    admin_token = getattr(request.state, "admin_token", None)
    if admin_token and "csrf_token" not in context:
        context["csrf_token"] = _make_csrf_token(admin_token)
    return templates.TemplateResponse(template, {"request": request, **context})


async def require_admin_cookie(request: Request, admin_access: str = Cookie(None)):
    if not admin_access:
        raise HTTPException(
            status_code=302,
            detail="Unauthorized",
            headers={"Location": "/dashboard/login?error=Unauthorized"}
        )
    try:
        payload = await auth_service._decode_token(admin_access)
        if payload.get("role") != "admin":
            raise HTTPException(
                status_code=302,
                detail="Access denied",
                headers={"Location": "/dashboard/login?error=Access+denied"}
            )
        request.state.admin_token = admin_access
        return payload
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(
            status_code=302,
            detail="Session expired",
            headers={"Location": "/dashboard/login?error=Session+expired"}
        )


async def require_csrf(
    request: Request,
    csrf_token: str = Form(None),
    admin_access: str = Cookie(None),
):
    if not admin_access:
        raise HTTPException(status_code=403, detail="Missing admin session")
    if not csrf_token or not _verify_csrf_token(admin_access, csrf_token):
        raise HTTPException(status_code=403, detail="Invalid CSRF token")

@router.get("/login")
async def login_page(request: Request, admin_access: str = Cookie(None)):
    if admin_access:
        try:
            payload = await auth_service._decode_token(admin_access)
            if payload.get("role") == "admin":
                return RedirectResponse("/dashboard/users")
        except Exception:
            pass
    return render_page("admin/login.html", request)

@router.post("/login")
async def admin_login(
    email: str = Form(...),
    password: str = Form(...),
    session: AsyncSession = Depends(get_async_session),
):
    user = await crud_user.get_user_by_email(session, email)
    if not user or not Hasher.verify_password(password, user.password):
        return RedirectResponse("/dashboard/login?error=Invalid+credentials", status_code=302)
    if user.role != "admin":
        return RedirectResponse("/dashboard/login?error=Access+denied", status_code=302)

    payload = {"id": user.id, "role": user.role}
    token = token_service.generate_token(payload=payload, expires_in_hours=6)
    res = RedirectResponse("/dashboard/users", status_code=302)
    res.set_cookie("admin_access", token, httponly=True, secure=True, samesite="strict", max_age=21600)
    return res

@router.post("/logout")
async def admin_logout(
    _csrf: None = Depends(require_csrf),
    _auth: dict = Depends(require_admin_cookie),
):
    res = RedirectResponse("/dashboard/login", status_code=302)
    res.delete_cookie("admin_access")
    return res

@router.get("/users")
async def users_page(
    request: Request, 
    auth= Depends(require_admin_cookie), 
    session: AsyncSession = Depends(get_async_session),
    success: str = None,
    error: str = None
    ):
    users = await crud_user.get_users(session=session)
    return render_page("admin/users.html", request, users=users, auth=auth, success=success, error=error)

@router.get("/users/{user_id}")
async def user_detail_page(
    request: Request,
    user_id: str,
    page: int = 1,
    limit: int = 5,
    sort: str = "created_at",
    order: str = "desc",
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    user = await crud_user.get_user_by_id(session, id=user_id)
    if not user:
        raise HTTPException(404, "User not found")

    offset = (page - 1) * limit

    valid_sorts = {
        "bmi": UserNutrition.bmi,
        "height": UserNutrition.height_cm,
        "weight": UserNutrition.weight_kg,
        "ideal": UserNutrition.ideal_weight_kg,
        "created_at": UserNutrition.created_at,
    }

    sort_column = valid_sorts.get(sort, UserNutrition.created_at)
    ordering = sort_column.desc() if order == "desc" else sort_column.asc()

    crud_nutrition = CRUDUserNutrition()
    nutritions = await crud_nutrition.get_list_sorted(
        session=session,
        user_id=user_id,
        limit=limit,
        offset=offset,
        ordering=ordering
    )

    for n in nutritions:
        n.created_at = n.created_at.astimezone(ZoneInfo("Asia/Jakarta"))

    all_nutritions_raw = await crud_nutrition.get_list(session, user_id)

    all_nutritions = [
        {
            "created_at": n.created_at.strftime("%Y-%m-%d"),
            "weight_kg": n.weight_kg,
            "ideal_weight_kg": n.ideal_weight_kg,
        }
        for n in all_nutritions_raw
    ]

    total_records = await crud_nutrition.count_by_user(session, user_id)
    total_pages = (total_records + limit - 1) // limit

    return render_page(
        "admin/user_detail.html",
        request,
        user=user,
        user_nutritions=nutritions,
        all_nutritions=all_nutritions,
        page=page,
        total_pages=total_pages,
        limit=limit,
        auth=auth,
    )

@router.post("/users/{user_id}/delete")
async def delete_user(
    request: Request,
    user_id: str,
    csrf_token: str = Form(...),
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    admin_token = getattr(request.state, "admin_token", None)
    if not admin_token or not _verify_csrf_token(admin_token, csrf_token):
        return RedirectResponse("/dashboard/users?error=Invalid+request", status_code=303)
    deleted = await crud_user.delete_user(session, user_id)
    if not deleted:
        return RedirectResponse("/dashboard/users?error=User+not+found", status_code=303)
    return RedirectResponse("/dashboard/users?success=User+deleted+successfully", status_code=303)


@router.get("/reset/password")
async def reset_password_page(request: Request):
    return render_page("admin/reset_password.html", request)

@router.post("/reset/password")
async def send_reset_link(
    request: Request,
    email: str = Form(...),
    session: AsyncSession = Depends(get_async_session),
):
    user = await crud_user.get_user_by_email(session, email)
    if user:
        payload = {"id": user.id, "email": user.email}
        token = token_service.generate_token(payload=payload, token_type="reset_password", expires_in_minutes=30)
        reset_link = f"https://sehatiapps.web.id/dashboard/reset/password/confirm?token={token}"
        try:
            email_client.send_password_reset_email(recipient=user.email, fullname=user.fullname, link=reset_link)
        except Exception as error:
            logger.error(error)
    return render_page("admin/reset_password.html", request, success="You will receive a verification email if the email address exists in our system.")

@router.get("/reset/password/confirm")
async def confirm_reset_password(request: Request, token: str):
    return render_page("admin/reset_confirm.html", request, token=token)

@router.post("/reset/password/confirm")
async def confirm_reset_password_post(
    request: Request,
    token: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    session: AsyncSession = Depends(get_async_session),
):
    if new_password != confirm_password:
        return render_page("admin/reset_confirm.html", request, token=token, error="Passwords do not match.")

    try:
        payload = await auth_service._decode_token(token)
        if payload.get("type") != "reset_password":
            raise ValueError("Invalid token type.")
        user_id = payload.get("id")
    except Exception:
        return render_page("admin/reset_password.html", request, error="Invalid or expired token.")

    user = await crud_user.get_user_admin_by_id(session, user_id)
    if not user:
        return render_page("admin/reset_password.html", request, error="User not found.")

    try:
        if re.fullmatch(settings.PASSWORD_REGEX, new_password) is None:
            raise ValueError("Password must be at least 8 chars, contain upper, lower, digit, and symbol.")
        user.password = Hasher.hash_password(new_password)
        await session.commit()
    except Exception:
        await session.rollback()
        return render_page("admin/reset_confirm.html", request, error="Database error. Try again later.")

    return render_page("admin/reset_success.html", request, message="Password successfully updated.")

@router.get("/recipes")
async def recipes_page(
    request: Request,
    page: int = 1,
    limit: int = 10,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    offset = (page - 1) * limit

    recipes = await crud_recipe.get_all(session, limit=limit, offset=offset)
    total = await crud_recipe.count(session)
    total_pages = (total + limit - 1) // limit

    return render_page(
        "admin/recipes.html",
        request,
        recipes=recipes,
        page=page,
        total_pages=total_pages,
        limit=limit,
        auth=auth,
    )

@router.get("/recipes/upload")
async def recipe_upload_page(
    request: Request,
    auth=Depends(require_admin_cookie)
):
    return render_page("admin/recipe_upload.html", request, auth=auth)

@router.post("/recipes/upload")
async def recipe_upload(
    request: Request,
    title: str = Form(...),
    description: str = Form(...),
    category: str = Form(None),
    file: UploadFile = Form(...),
    image: UploadFile = Form(None),
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
):
    file_url = await save_upload_with_uuid(file, folder="recipe")
    image_url = await save_upload_with_uuid(image, folder="recipe")

    await crud_recipe.create(
        session=session,
        data={
            "title": title,
            "description": description,
            "file_url": f"/media/recipe/{file_url}",
            "image_url": f"/media/recipe/{image_url}",
            "category": category,
        }
    )

    return RedirectResponse("/dashboard/recipes", status_code=302)

@router.get("/leaderboard")
async def leaderboard_page(
    request: Request,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session)
):
    wallets = await crud_wallet.get_all(session=session)
    data = [
        {
            "nickname": w.user.nickname,
            "fullname": w.user.fullname,
            "achievement_points": w.achievement_points,
            "credit_points": w.credit_points,
        }
        for w in wallets
    ]

    return render_page("admin/leaderboard.html", request, leaderboard=data, auth=auth)

@router.get("/merchandise/upload")
async def merchandise_upload_page(
    request: Request,
    auth=Depends(require_admin_cookie)
):
    return render_page("admin/merchandise_upload.html", request, auth=auth)

@router.post("/merchandise/upload")
async def merchandise_upload(
    request: Request,
    name: str = Form(...),
    description: str = Form(None),
    price_points: int = Form(...),
    stock: int = Form(...),
    image: UploadFile = Form(...),
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
):
    image_url = await save_upload_with_uuid(image, folder="merchandise")

    await crud_merch.create(
        session=session,
        data={
            "name": name,
            "description": description,
            "price_points": price_points,
            "stock": stock,
            "image_url": f"/media/merchandise/{image_url}",
        }
    )

    return RedirectResponse("/dashboard/merchandise", status_code=302)

@router.get("/merchandise")
async def merchandise_page(
    request: Request,
    page: int = 1,
    limit: int = 10,
    success: str = None,
    error: str = None,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    offset = (page - 1) * limit

    merch = await crud_merch.get_all(session, limit=limit, offset=offset)
    total = await crud_merch.count(session)
    total_pages = (total + limit - 1) // limit

    return render_page(
        "admin/merchandise.html",
        request,
        merch=merch,
        page=page,
        total_pages=total_pages,
        limit=limit,
        auth=auth,
        success=success,
        error=error,
    )

@router.post("/merchandise/update/{merch_id}")
async def update_merchandise(
    merch_id: str,
    name: str = Form(...),
    price_points: int = Form(...),
    stock: int = Form(...),
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie)
):
    merchandise = await crud_merch.get_by_id(session=session, id=merch_id)
    if not merchandise:
        return RedirectResponse(
            "/dashboard/merchandise?error=Merchandise not found",
            status_code=302
        )

    merchandise.name = name
    merchandise.price_points = price_points
    merchandise.stock = stock

    await session.commit()
    await session.refresh(merchandise)

    return RedirectResponse(
        "/dashboard/merchandise?success=Updated successfully",
        status_code=302
    )

@router.get("/merchandise/claims")
async def merchandise_claims_page(
    request: Request,
    page: int = 1,
    limit: int = 10,
    error: str = None,
    success: str = None,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    offset = (page - 1) * limit

    claims = await crud_merch_claim.get_list(session=session, limit=limit, offset=offset)
    total = await crud_merch_claim.count(session=session)
    total_pages = (total + limit - 1) // limit

    return render_page(
        "admin/merch_claims.html",
        request,
        claims=claims,
        page=page,
        total_pages=total_pages,
        limit=limit,
        auth=auth,
        success=success,
        error=error
    )

@router.post("/merchandise/claims/{claim_id}/approve")
async def approve_claim(claim_id: str, session: AsyncSession = Depends(get_async_session), auth=Depends(require_admin_cookie)):
    merchandise_claim = await crud_merch_claim.get_by_id(id=claim_id, session=session)
    merchandise = merchandise_claim.merchandise

    if merchandise.stock <= 0:
        return RedirectResponse(
            "/dashboard/merchandise/claims?error=Stock+unavailable+Please+add+stock+on+the+merchandise+list+page+or+reject+this+claim",
            status_code=302
        )
    
    user = await crud_user.get_user_by_id(session=session, id=merchandise_claim.user_id)
    await crud_merch_claim.update_status(session, claim_id=claim_id, status="approved")
    await crud_merch.update_stock(session=session, id=merchandise_claim.merchandise_id)
    await redeem_merchandise_points(session=session, user_id=user.id, merchandise_id=merchandise.id)

    try:
        email_client.send_approve_claim_marchandise(
            recipient=user.email,
            context={
                "fullname": user.fullname,
                "merchandise_name": merchandise_claim.merchandise.name,
                "year": datetime.now().year
            }
        )
    except Exception as error:
        logger.error(error)

    return RedirectResponse("/dashboard/merchandise/claims?success=Approved", status_code=302)

@router.post("/merchandise/claims/{claim_id}/reject")
async def reject_claim(claim_id: str, session: AsyncSession = Depends(get_async_session), auth=Depends(require_admin_cookie)):
    merchandise_claim = await crud_merch_claim.get_by_id(id=claim_id, session=session)

    if not merchandise_claim:
        return RedirectResponse("/dashboard/merchandise/claims?error=Claim+not+found", status_code=302)

    if merchandise_claim.status != "pending":
        return RedirectResponse("/dashboard/merchandise/claims?error=Claim+already+processed", status_code=302)

    user = await crud_user.get_user_by_id(session=session, id=merchandise_claim.user_id)
    await crud_merch_claim.update_status(session, claim_id=claim_id, status="rejected")

    try:
        email_client.send_rejected_claim_marchandise(
            recipient=user.email,
            context={
                "fullname": user.fullname,
                "merchandise_name": merchandise_claim.merchandise.name,
                "year": datetime.now().year,
            }
        )
    except Exception as error:
        logger.error(error)

    return RedirectResponse("/dashboard/merchandise/claims?success=Claim+has+been+rejected", status_code=302)

@router.get("/transactions")
async def point_transactions_page(
    request: Request,
    name: str = None,
    page: int = 1,
    limit: int = 20,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session)
):
    offset = (page - 1) * limit

    transactions = await crud_transaction.get_history(session, name=name, limit=limit, offset=offset)
    for t in transactions:
        t.created_at = t.created_at.astimezone(ZoneInfo("Asia/Jakarta"))
    total = await crud_transaction.count(session, name=name)
    total_pages = (total + limit - 1) // limit

    page_numbers = []
    if total_pages <= 7:
        page_numbers = list(range(1, total_pages + 1))
    else:
        page_numbers.append(1)
        
        if page <= 3:
            page_numbers.extend([2, 3, 4])
            page_numbers.append("...")
            page_numbers.append(total_pages)
        elif page >= total_pages - 2:
            page_numbers.append("...")
            page_numbers.extend([total_pages - 3, total_pages - 2, total_pages - 1, total_pages])
        else:
            page_numbers.append("...")
            page_numbers.extend([page - 1, page, page + 1])
            page_numbers.append("...")
            page_numbers.append(total_pages)

    return render_page(
        "admin/point_transactions.html",
        request,
        transactions=transactions,
        name=name,
        page=page,
        limit=limit,
        total_pages=total_pages,
        page_numbers=page_numbers,
        auth=auth,
    )

@router.get("/transactions/export")
async def export_point_transactions(
    request: Request,
    name: str = None,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session)
):
    transactions = await crud_transaction.get_all_for_export(session, name=name)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Point Transactions"
    
    headers = [
        "User Name",
        "Wallet",
        "Type",
        "Category",
        "Points",
        "Balance",
        "Date"
    ]
    
    header_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for row_num, tx in enumerate(transactions, 2):
        created_at = tx.created_at.astimezone(ZoneInfo("Asia/Jakarta"))
        
        data = [
            tx.user.fullname,
            tx.wallet.value,
            tx.tx_type.value,
            tx.category_code.value,
            tx.delta,
            tx.balance_after or 0,
            created_at.strftime("%Y-%m-%d %H:%M:%S")
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            if col_num == 5:
                if tx.tx_type.value == "earn":
                    cell.font = Font(color="008000", bold=True)
                elif tx.tx_type.value == "spend":
                    cell.font = Font(color="FF0000", bold=True)
            
            if col_num in [7, 8]:
                cell.alignment = Alignment(horizontal='right')
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"point_transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    if name:
        filename = f"point_transactions_{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/games")
async def games_page(
    request: Request,
    page: int = 1,
    limit: int = 10,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    offset = (page - 1) * limit

    games = await crud_games.get_all(session, limit=limit, offset=offset)
    total = await crud_games.count(session)
    total_pages = (total + limit - 1) // limit

    return render_page(
        "admin/games.html",
        request,
        games=games,
        page=page,
        total_pages=total_pages,
        limit=limit,
        auth=auth,
    )

@router.get("/games/create")
async def game_create_page(request: Request, auth=Depends(require_admin_cookie)):
    return render_page("admin/game_create.html", request, auth=auth)

@router.post("/games/create")
async def create_game(
    request: Request,
    name: str = Form(...),
    description: str = Form(None),
    url: str = Form(...),
    price_points: int = Form(...),
    image: UploadFile = Form(...),
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    image_url = await save_upload_with_uuid(image, folder="games")

    await crud_games.create(
        session=session,
        data={
            "name": name,
            "description": description,
            "url": url,
            "price_points": price_points,
            "image_url": f"/media/games/{image_url}",
        }
    )

    return RedirectResponse("/dashboard/games", status_code=302)

@router.get("/games/{game_id}/view")
async def view_game(
    request: Request,
    game_id: str,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session)
):
    game = await crud_games.get_by_id(session=session, id=game_id)

    if not game:
        raise HTTPException(status_code=404, detail="Game not found")

    return render_page("admin/game.html", request, game=game, auth=auth)

@router.get("/foods")
async def foods_page(
    request: Request,
    name: str = None,
    page: int = 1,
    limit: int = 10,
    success: str = None,
    error: str = None,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    offset = (page - 1) * limit

    foods = await crud_food.get_all(session, name=name, limit=limit, offset=offset)
    total = await crud_food.count(session, name=name)
    total_pages = (total + limit - 1) // limit

    return templates.TemplateResponse("admin/foods.html", {
        "request": request,
        "foods": foods,
        "name": name,
        "page": page,
        "total_pages": total_pages,
        "limit": limit,
        "auth": auth,
        "success": success,
        "error": error,
    })

@router.get("/foods/create")
async def food_create_page(
    request: Request,
    auth=Depends(require_admin_cookie)
):
    return render_page("admin/food_create.html", request, auth=auth)

@router.post("/foods/create")
async def create_food(
    request: Request,
    name: str = Form(...),
    calories: int = Form(...),
    unit: str = Form("kcal"),
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
):
    # Cek jika food sudah ada
    existing = await crud_food.get_by_name(session=session, name=name)
    if existing:
        return RedirectResponse(
            "/dashboard/foods/create?error=Food+already+exists",
            status_code=302
        )

    await crud_food.create(
        session=session,
        data={
            "name": name,
            "calories": calories,
            "unit": unit or "kcal",
        }
    )

    return RedirectResponse("/dashboard/foods?success=Created+successfully", status_code=302)

@router.post("/foods/update/{food_id}")
async def update_food(
    food_id: str,
    name: str = Form(...),
    calories: int = Form(...),
    unit: str = Form("kcal"),
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie)
):
    food = await crud_food.get_by_id(session=session, food_id=food_id)
    if not food:
        return RedirectResponse("/dashboard/foods?error=Food+not+found", status_code=302)
    await crud_food.update(session=session, food_id=food_id, data={"name":name, "calories":calories, "unit":unit})
    return RedirectResponse("/dashboard/foods?success=Updated+successfully", status_code=302)

@router.post("/foods/delete/{food_id}")
async def delete_food(
    food_id: str,
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
):
    food = await crud_food.get_by_id(session=session, food_id=food_id)
    if not food:
        return RedirectResponse("/dashboard/foods?error=Food+not+found", status_code=302)

    await crud_food.delete(session=session, food_id=food_id)

    return RedirectResponse("/dashboard/foods?success=Deleted", status_code=302)

@router.get("/export/health-data")
async def export_health_data_excel(
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
):
    try:
        exporter = HealthDataExcelExporter(session)
        excel_file = await exporter.generate_excel()
        filename = f"health_research_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except Exception as e:
        print(f"Error exporting Excel: {str(e)}")
        return RedirectResponse(
            "/dashboard/users?error=Failed+to+export+data", 
            status_code=302
        )

@router.get("/appointments")
async def appointments_page(
    request: Request,
    username: str = None,
    status: str = None,
    page: int = 1,
    limit: int = 10,
    success: str = None,
    error: str = None,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    offset = (page - 1) * limit

    appointments = await crud_appointment.get_all(
        session=session,
        username=username,
        status=status,
        limit=limit,
        offset=offset
    )
    
    total = await crud_appointment.count(session=session, username=username, status=status)
    total_pages = (total + limit - 1) // limit

    return render_page(
        "admin/appointments.html",
        request,
        appointments=appointments,
        username=username,
        status=status,
        page=page,
        total_pages=total_pages,
        limit=limit,
        auth=auth,
        success=success,
        error=error,
    )


@router.post("/appointments/update/{appointment_id}/{new_status}")
async def update_appointment_status(
    appointment_id: str,
    new_status: str,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    try:
        await crud_appointment.update_status(
            session=session,
            id=appointment_id,
            status=new_status
        )
        return RedirectResponse(
            url="/dashboard/appointments?success=Status updated successfully",
            status_code=303
        )
    except Exception as e:
        logger.error(f"Error updating appointment: {str(e)}")
        return RedirectResponse(
            url="/dashboard/appointments?error=Failed to update status",
            status_code=303
        )


@router.post("/appointments/delete/{appointment_id}")
async def delete_appointment(
    appointment_id: str,
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
    _csrf: None = Depends(require_csrf),
):
    deleted = await crud_appointment.delete(session=session, id=appointment_id)
    if not deleted:
        return RedirectResponse(
            url="/dashboard/appointments?error=Appointment+not+found",
            status_code=303,
        )
    return RedirectResponse(
        url="/dashboard/appointments?success=Appointment+deleted",
        status_code=303,
    )

@router.get("/videos")
async def videos_page(
    request: Request,
    title: Optional[str] = None,
    page: int = 1,
    limit: int = 10,
    success: Optional[str] = None,
    error: Optional[str] = None,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    offset = (page - 1) * limit
    videos = await crud_video.get_all_admin(session, limit=limit, offset=offset, title=title)
    total = await crud_video.count(session, title=title)
    total_pages = (total + limit - 1) // limit
    return render_page(
        "admin/videos.html",
        request,
        videos=videos,
        title=title,
        page=page,
        total_pages=total_pages,
        limit=limit,
        auth=auth,
        success=success,
        error=error,
    )


@router.get("/videos/create")
async def video_create_page(request: Request, auth=Depends(require_admin_cookie)):
    return render_page("admin/video_create.html", request, auth=auth)


@router.post("/videos/create")
async def create_video(
    request: Request,
    title: str = Form(...),
    description: str = Form(None),
    youtube_url: str = Form(...),
    thumbnail: str = Form(None),
    category: str = Form(None),
    duration_seconds: str = Form(None),
    is_active: str = Form(None),
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
    _csrf: None = Depends(require_csrf),
):
    try:
        dur = int(duration_seconds) if duration_seconds and duration_seconds.strip() else 0
    except ValueError:
        dur = 0
    video = Video(
        title=title,
        description=description or None,
        youtube_url=youtube_url,
        thumbnail=thumbnail or None,
        category=category or None,
        duration_seconds=dur,
        is_active=(is_active == "on"),
    )
    await crud_video.create(session=session, video=video)
    return RedirectResponse("/dashboard/videos?success=Video+created+successfully", status_code=302)


@router.post("/videos/delete/{video_id}")
async def delete_video(
    video_id: str,
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
    _csrf: None = Depends(require_csrf),
):
    deleted = await crud_video.soft_delete(session=session, id=video_id)
    if not deleted:
        return RedirectResponse("/dashboard/videos?error=Video+not+found", status_code=302)
    return RedirectResponse("/dashboard/videos?success=Video+deleted", status_code=302)


@router.post("/videos/toggle/{video_id}")
async def toggle_video(
    video_id: str,
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
    _csrf: None = Depends(require_csrf),
):
    video = await crud_video.get_by_id(id=video_id, session=session)
    if not video:
        return RedirectResponse("/dashboard/videos?error=Video+not+found", status_code=302)
    await crud_video.update(session=session, id=video_id, data={"is_active": not video.is_active})
    status = "activated" if not video.is_active else "deactivated"
    return RedirectResponse(f"/dashboard/videos?success=Video+{status}", status_code=302)


@router.get("/professionals")
async def professionals_page(
    request: Request,
    page: int = 1,
    limit: int = 10,
    success: Optional[str] = None,
    error: Optional[str] = None,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    offset = (page - 1) * limit
    professionals = await crud_professional.list_all(session, limit=limit, offset=offset)
    total = await crud_professional.count(session)
    total_pages = (total + limit - 1) // limit
    return render_page(
        "admin/professionals.html",
        request,
        professionals=professionals,
        page=page,
        total_pages=total_pages,
        limit=limit,
        auth=auth,
        success=success,
        error=error,
    )


@router.get("/professionals/create")
async def professional_create_page(request: Request, auth=Depends(require_admin_cookie)):
    return render_page("admin/professional_create.html", request, auth=auth)


@router.post("/professionals/create")
async def create_professional(
    request: Request,
    fullname: str = Form(...),
    email: str = Form(...),
    phone_number: str = Form(None),
    specialization: str = Form(...),
    bio: str = Form(None),
    hour_start: str = Form("09:00"),
    hour_end: str = Form("17:00"),
    day_monday: str = Form(None),
    day_tuesday: str = Form(None),
    day_wednesday: str = Form(None),
    day_thursday: str = Form(None),
    day_friday: str = Form(None),
    day_saturday: str = Form(None),
    day_sunday: str = Form(None),
    is_active: str = Form(None),
    picture: UploadFile = File(None),
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
    _csrf: None = Depends(require_csrf),
):
    available_days = {
        "monday": day_monday == "on",
        "tuesday": day_tuesday == "on",
        "wednesday": day_wednesday == "on",
        "thursday": day_thursday == "on",
        "friday": day_friday == "on",
        "saturday": day_saturday == "on",
        "sunday": day_sunday == "on",
    }
    available_hours = {"start": hour_start or "09:00", "end": hour_end or "17:00"}

    picture_url = None
    if picture and picture.filename:
        try:
            filename = await save_upload_with_uuid(picture, folder="avatars")
            picture_url = f"/media/avatars/{filename}"
        except Exception as e:
            logger.error(f"Picture upload failed: {e}")

    try:
        await crud_professional.create(
            session=session,
            data={
                "fullname": fullname,
                "email": email,
                "phone_number": phone_number or None,
                "specialization": specialization,
                "bio": bio or None,
                "picture": picture_url,
                "available_days": available_days,
                "available_hours": available_hours,
                "is_active": is_active == "on",
            },
        )
    except Exception as e:
        logger.error(f"Create professional error: {e}")
        return RedirectResponse(
            "/dashboard/professionals/create?error=Email+already+exists+or+invalid+data",
            status_code=302,
        )
    return RedirectResponse(
        "/dashboard/professionals?success=Professional+created+successfully", status_code=302
    )


@router.get("/professionals/{professional_id}/edit")
async def professional_edit_page(
    request: Request,
    professional_id: str,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    professional = await crud_professional.get_by_id(session=session, id=professional_id)
    if not professional:
        return RedirectResponse("/dashboard/professionals?error=Professional+not+found", status_code=302)
    return render_page("admin/professional_create.html", request, auth=auth, professional=professional)


@router.post("/professionals/{professional_id}/update")
async def update_professional(
    request: Request,
    professional_id: str,
    fullname: str = Form(...),
    email: str = Form(...),
    phone_number: str = Form(None),
    specialization: str = Form(...),
    bio: str = Form(None),
    hour_start: str = Form("09:00"),
    hour_end: str = Form("17:00"),
    day_monday: str = Form(None),
    day_tuesday: str = Form(None),
    day_wednesday: str = Form(None),
    day_thursday: str = Form(None),
    day_friday: str = Form(None),
    day_saturday: str = Form(None),
    day_sunday: str = Form(None),
    is_active: str = Form(None),
    picture: UploadFile = File(None),
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
    _csrf: None = Depends(require_csrf),
):
    available_days = {
        "monday": day_monday == "on",
        "tuesday": day_tuesday == "on",
        "wednesday": day_wednesday == "on",
        "thursday": day_thursday == "on",
        "friday": day_friday == "on",
        "saturday": day_saturday == "on",
        "sunday": day_sunday == "on",
    }
    available_hours = {"start": hour_start or "09:00", "end": hour_end or "17:00"}

    data = {
        "fullname": fullname,
        "email": email,
        "phone_number": phone_number or None,
        "specialization": specialization,
        "bio": bio or None,
        "available_days": available_days,
        "available_hours": available_hours,
        "is_active": is_active == "on",
    }

    if picture and picture.filename:
        try:
            filename = await save_upload_with_uuid(picture, folder="avatars")
            data["picture"] = f"/media/avatars/{filename}"
        except Exception as e:
            logger.error(f"Picture upload failed: {e}")

    updated = await crud_professional.update(session=session, id=professional_id, data=data)
    if not updated:
        return RedirectResponse("/dashboard/professionals?error=Professional+not+found", status_code=302)
    return RedirectResponse(
        "/dashboard/professionals?success=Professional+updated+successfully", status_code=302
    )


@router.post("/professionals/{professional_id}/delete")
async def delete_professional(
    professional_id: str,
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
    _csrf: None = Depends(require_csrf),
):
    deleted = await crud_professional.delete(session=session, id=professional_id)
    if not deleted:
        return RedirectResponse("/dashboard/professionals?error=Professional+not+found", status_code=302)
    return RedirectResponse(
        "/dashboard/professionals?success=Professional+deleted", status_code=302
    )


# ──────────────────────────── BLAST EMAIL ────────────────────────────────────

@router.get("/blast")
async def blast_page(
    request: Request,
    success: Optional[str] = None,
    error: Optional[str] = None,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    users = await crud_user.get_active_users(session)
    blast_history = await crud_blast_log.get_all(session, limit=20)
    return render_page(
        "admin/blast.html",
        request,
        users=users,
        blast_history=blast_history,
        auth=auth,
        success=success,
        error=error,
    )


async def _send_and_log_blast(blast_id: str, recipients: list, subject: str, body_html: str, cc: Optional[str]):
    _blast_progress[blast_id] = {"sent": 0, "failed": 0, "total": len(recipients)}

    def on_progress(sent: int, failed: int):
        _blast_progress[blast_id]["sent"] = sent
        _blast_progress[blast_id]["failed"] = failed

    async_gen = get_async_session()
    session = await async_gen.__anext__()
    try:
        result = await asyncio.to_thread(email_client.send_blast, recipients, subject, body_html, cc, on_progress)
        status = "completed" if result["failed"] == 0 else ("partial_failed" if result["sent"] > 0 else "failed")
        await crud_blast_log.update_status(session, blast_id, status, result["sent"], result["failed"], result.get("failed_emails", {}))
    except Exception as e:
        logger.error(f"Background blast task failed: {e}")
        try:
            await crud_blast_log.update_status(session, blast_id, "failed", 0, len(recipients), {})
        except Exception:
            pass
    finally:
        _blast_progress.pop(blast_id, None)
        await session.close()


@router.post("/blast")
async def send_blast(
    request: Request,
    background_tasks: BackgroundTasks,
    subject: str = Form(...),
    body: str = Form(...),
    cc: str = Form(None),
    recipient_mode: str = Form("all"),
    selected_users: Optional[FormList[str]] = Form(default=None),
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
    _csrf: None = Depends(require_csrf),
):
    active_users = await crud_user.get_active_users(session)

    if recipient_mode == "specific":
        ids = set(selected_users or [])
        if not ids:
            if request.headers.get("X-Requested-With") == "fetch":
                return {"status": "error", "message": "Please select at least one recipient"}
            return RedirectResponse("/dashboard/blast?error=Please+select+at+least+one+recipient", status_code=302)
        recipients = [u.email for u in active_users if u.id in ids]
    else:
        recipients = [u.email for u in active_users]

    if not recipients:
        if request.headers.get("X-Requested-With") == "fetch":
            return {"status": "error", "message": "No valid recipients found"}
        return RedirectResponse("/dashboard/blast?error=No+valid+recipients+found", status_code=302)

    blast_html = email_client._render(
        "emails/blast.html",
        {"subject": subject, "body": body, "year": datetime.now().year},
    )
    cc_clean = cc.strip() if cc and cc.strip() else None

    blast_log = await crud_blast_log.create(session, subject, body, recipients, cc_clean, auth.get("id"))
    background_tasks.add_task(_send_and_log_blast, blast_log.id, recipients, subject, blast_html, cc_clean)

    count = len(recipients)
    if request.headers.get("X-Requested-With") == "fetch":
        return {"status": "success", "blast_id": blast_log.id, "message": f"Email is being sent to {count} recipients"}
    return RedirectResponse(
        f"/dashboard/blast?success=Email+is+being+sent+to+{count}+recipients",
        status_code=302,
    )


@router.get("/blast/{blast_id}/detail")
async def blast_detail(
    blast_id: str,
    request: Request,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    blast = await crud_blast_log.get_by_id(session, blast_id)
    if not blast:
        raise HTTPException(status_code=404, detail="Blast not found")
    return render_page(
        "admin/blast_detail.html",
        request,
        blast=blast,
        auth=auth,
    )


@router.get("/api/blast/{blast_id}/status")
async def get_blast_status(
    blast_id: str,
    auth=Depends(require_admin_cookie),
    session: AsyncSession = Depends(get_async_session),
):
    # Real-time in-memory progress (while task is running)
    if blast_id in _blast_progress:
        prog = _blast_progress[blast_id]
        return {
            "id": blast_id,
            "status": "pending",
            "sent_count": prog["sent"],
            "failed_count": prog["failed"],
            "total_count": prog["total"],
        }
    # Fallback to DB (task finished or historical)
    blast = await crud_blast_log.get_by_id(session, blast_id)
    if not blast:
        raise HTTPException(status_code=404, detail="Blast not found")
    return {
        "id": blast.id,
        "status": blast.status,
        "sent_count": blast.sent_count,
        "failed_count": blast.failed_count,
        "total_count": len(blast.recipients),
    }


@router.post("/blast/{blast_id}/retry")
async def retry_blast(
    blast_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    session: AsyncSession = Depends(get_async_session),
    auth=Depends(require_admin_cookie),
    _csrf: None = Depends(require_csrf),
):
    blast = await crud_blast_log.get_by_id(session, blast_id)
    if not blast:
        return RedirectResponse("/dashboard/blast?error=Blast+not+found", status_code=302)

    blast_html = email_client._render(
        "emails/blast.html",
        {"subject": blast.subject, "body": blast.body_html, "year": datetime.now().year},
    )
    background_tasks.add_task(_send_and_log_blast, blast.id, blast.recipients, blast.subject, blast_html, blast.cc)

    count = len(blast.recipients)
    return RedirectResponse(
        f"/dashboard/blast?success=Retrying+email+to+{count}+recipients",
        status_code=302,
    )